"""複数店舗のExcel売上ファイルを集約し、本部用の月次レポートExcelを生成する.

使い方:
    pip install openpyxl
    python monthly_report.py                 # サンプル入力(input/)で実行
    python monthly_report.py --input-dir /path --output report.xlsx
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl が必要です。`pip install openpyxl` を実行してください。")
    raise


HERE = Path(__file__).resolve().parent


@dataclass
class Sale:
    store: str
    date: str
    product: str
    category: str
    quantity: int
    unit_price: int

    @property
    def revenue(self) -> int:
        return self.quantity * self.unit_price


# --------------- 入力読み込み ---------------


def load_store_excels(input_dir: Path, sheet_name: str | None = None) -> list[Sale]:
    sales: list[Sale] = []
    files = sorted(input_dir.glob("store_*.xlsx"))
    if not files:
        raise FileNotFoundError(f"{input_dir} に store_*.xlsx が見つかりません")

    for fp in files:
        store = fp.stem.removeprefix("store_")
        wb = load_workbook(fp, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(h).strip() if h else "" for h in rows[0]]
        idx = {col: header.index(col) for col in ("date", "product", "category", "quantity", "unit_price") if col in header}
        if len(idx) < 5:
            print(f"[WARN] {fp.name}: 必要な列が揃っていません. skip")
            continue
        for r in rows[1:]:
            try:
                date_v = r[idx["date"]]
                date_str = (
                    date_v.strftime("%Y-%m-%d") if isinstance(date_v, datetime) else str(date_v)
                )
                sales.append(
                    Sale(
                        store=store,
                        date=date_str,
                        product=str(r[idx["product"]]),
                        category=str(r[idx["category"]]),
                        quantity=int(r[idx["quantity"]]),
                        unit_price=int(r[idx["unit_price"]]),
                    )
                )
            except (ValueError, TypeError, AttributeError, IndexError):
                continue
        wb.close()
    return sales


# --------------- 集計 ---------------


def aggregate(sales: list[Sale]) -> dict:
    by_store: dict[str, dict] = defaultdict(lambda: {"revenue": 0, "qty": 0, "count": 0})
    by_product: dict[tuple[str, str], dict] = defaultdict(lambda: {"revenue": 0, "qty": 0})
    by_category: dict[str, int] = defaultdict(int)
    by_date: dict[str, int] = defaultdict(int)

    for s in sales:
        by_store[s.store]["revenue"] += s.revenue
        by_store[s.store]["qty"] += s.quantity
        by_store[s.store]["count"] += 1
        by_product[(s.product, s.category)]["revenue"] += s.revenue
        by_product[(s.product, s.category)]["qty"] += s.quantity
        by_category[s.category] += s.revenue
        by_date[s.date] += s.revenue

    total_rev = sum(s.revenue for s in sales)
    total_qty = sum(s.quantity for s in sales)
    return {
        "kpi": {
            "total_revenue": total_rev,
            "total_count": len(sales),
            "total_qty": total_qty,
            "avg_unit_price": (total_rev // total_qty) if total_qty else 0,
            "first_date": min((s.date for s in sales), default=""),
            "last_date": max((s.date for s in sales), default=""),
            "store_count": len(by_store),
        },
        "store": dict(by_store),
        "product": [
            {"product": p, "category": c, **vals}
            for (p, c), vals in by_product.items()
        ],
        "category": dict(by_category),
        "date": dict(by_date),
    }


# --------------- スタイル定義 ---------------


HEADER_FILL = PatternFill("solid", fgColor="2563EB")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="メイリオ")
KPI_FILL = PatternFill("solid", fgColor="DBEAFE")
KPI_FONT = Font(bold=True, size=14, name="メイリオ")
BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
BODY_FONT = Font(name="メイリオ", size=10)


def style_header(cell) -> None:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def style_body(cell, num: bool = False) -> None:
    cell.font = BODY_FONT
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="right" if num else "left")
    if num:
        cell.number_format = "#,##0"


def autofit(ws, padding: int = 2) -> None:
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + padding, 10), 40)


# --------------- レポート出力 ---------------


def write_report(out_path: Path, sales: list[Sale], agg: dict) -> None:
    wb = Workbook()
    _write_summary(wb.active, agg)
    wb.active.title = "サマリ"

    # 店舗別シート
    for store_name in sorted(agg["store"].keys()):
        ws = wb.create_sheet(f"店舗_{store_name}")
        store_sales = [s for s in sales if s.store == store_name]
        _write_store_sheet(ws, store_name, store_sales)

    # 商品別シート
    ws = wb.create_sheet("商品別")
    _write_product_sheet(ws, agg["product"])

    # 元データシート
    ws = wb.create_sheet("明細")
    _write_raw_sheet(ws, sales)

    wb.save(out_path)


def _write_summary(ws, agg: dict) -> None:
    kpi = agg["kpi"]
    ws["A1"] = "📊 月次売上レポート"
    ws["A1"].font = Font(bold=True, size=18, name="メイリオ")
    ws.merge_cells("A1:F1")

    ws["A2"] = f"期間: {kpi['first_date']} 〜 {kpi['last_date']} / 店舗数: {kpi['store_count']}"
    ws["A2"].font = Font(italic=True, color="6B7280", name="メイリオ")
    ws.merge_cells("A2:F2")

    ws["A4"] = "総売上"
    ws["B4"] = kpi["total_revenue"]
    ws["C4"] = "取引件数"
    ws["D4"] = kpi["total_count"]
    ws["E4"] = "平均単価"
    ws["F4"] = kpi["avg_unit_price"]

    for col in "ACE":
        ws[f"{col}4"].font = Font(bold=True, color="6B7280", name="メイリオ", size=10)
    for col in "BDF":
        ws[f"{col}4"].fill = KPI_FILL
        ws[f"{col}4"].font = KPI_FONT
        ws[f"{col}4"].number_format = "¥#,##0"

    # 店舗別ランキング
    ws["A6"] = "🏆 店舗別売上ランキング"
    ws["A6"].font = Font(bold=True, size=12, name="メイリオ")
    ws.merge_cells("A6:F6")

    headers = ["順位", "店舗", "売上", "取引件数", "平均単価", "売上構成比"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=7, column=i, value=h)
        style_header(c)

    sorted_stores = sorted(agg["store"].items(), key=lambda x: x[1]["revenue"], reverse=True)
    total = kpi["total_revenue"] or 1
    for rank, (store, vals) in enumerate(sorted_stores, 1):
        row = 7 + rank
        avg_unit = vals["revenue"] // vals["qty"] if vals["qty"] else 0
        share = vals["revenue"] / total * 100
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=store)
        ws.cell(row=row, column=3, value=vals["revenue"])
        ws.cell(row=row, column=4, value=vals["count"])
        ws.cell(row=row, column=5, value=avg_unit)
        ws.cell(row=row, column=6, value=share)
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            style_body(cell, num=col != 2)
        ws.cell(row=row, column=3).number_format = "¥#,##0"
        ws.cell(row=row, column=5).number_format = "¥#,##0"
        ws.cell(row=row, column=6).number_format = "0.00%"
        ws.cell(row=row, column=6).value = share / 100

    # カテゴリ別
    cat_start = 7 + len(sorted_stores) + 3
    ws.cell(row=cat_start, column=1, value="📂 カテゴリ別売上").font = Font(bold=True, size=12, name="メイリオ")
    ws.merge_cells(start_row=cat_start, start_column=1, end_row=cat_start, end_column=3)
    for i, h in enumerate(["カテゴリ", "売上", "構成比"], 1):
        style_header(ws.cell(row=cat_start + 1, column=i, value=h))
    sorted_cats = sorted(agg["category"].items(), key=lambda x: x[1], reverse=True)
    for i, (cat, rev) in enumerate(sorted_cats):
        r = cat_start + 2 + i
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=rev)
        ws.cell(row=r, column=3, value=rev / total)
        for col in range(1, 4):
            style_body(ws.cell(row=r, column=col), num=col != 1)
        ws.cell(row=r, column=2).number_format = "¥#,##0"
        ws.cell(row=r, column=3).number_format = "0.00%"

    autofit(ws)


def _write_store_sheet(ws, store_name: str, sales: list[Sale]) -> None:
    ws["A1"] = f"🏪 {store_name} 店"
    ws["A1"].font = Font(bold=True, size=16, name="メイリオ")
    ws.merge_cells("A1:F1")

    rev = sum(s.revenue for s in sales)
    qty = sum(s.quantity for s in sales)
    ws["A2"] = f"売上 ¥{rev:,} / 取引 {len(sales):,}件 / 数量 {qty:,}"
    ws["A2"].font = Font(italic=True, color="6B7280", name="メイリオ")
    ws.merge_cells("A2:F2")

    headers = ["日付", "商品", "カテゴリ", "数量", "単価", "売上"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(row=4, column=i, value=h))
    for i, s in enumerate(sorted(sales, key=lambda x: x.date), 1):
        r = 4 + i
        for col, v in enumerate(
            [s.date, s.product, s.category, s.quantity, s.unit_price, s.revenue], 1
        ):
            cell = ws.cell(row=r, column=col, value=v)
            style_body(cell, num=col >= 4)
            if col >= 5:
                cell.number_format = "¥#,##0"
    ws.freeze_panes = "A5"
    autofit(ws)


def _write_product_sheet(ws, product_data: list[dict]) -> None:
    ws["A1"] = "🛍️ 商品別売上ランキング"
    ws["A1"].font = Font(bold=True, size=16, name="メイリオ")
    ws.merge_cells("A1:E1")

    headers = ["順位", "商品", "カテゴリ", "数量", "売上"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=i, value=h))
    sorted_prod = sorted(product_data, key=lambda x: x["revenue"], reverse=True)
    for rank, p in enumerate(sorted_prod, 1):
        r = 3 + rank
        ws.cell(row=r, column=1, value=rank)
        ws.cell(row=r, column=2, value=p["product"])
        ws.cell(row=r, column=3, value=p["category"])
        ws.cell(row=r, column=4, value=p["qty"])
        ws.cell(row=r, column=5, value=p["revenue"])
        for col in range(1, 6):
            style_body(ws.cell(row=r, column=col), num=col not in (2, 3))
        ws.cell(row=r, column=5).number_format = "¥#,##0"
    ws.freeze_panes = "A4"
    autofit(ws)


def _write_raw_sheet(ws, sales: list[Sale]) -> None:
    headers = ["店舗", "日付", "商品", "カテゴリ", "数量", "単価", "売上"]
    for i, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=i, value=h))
    for i, s in enumerate(sales, 1):
        r = 1 + i
        for col, v in enumerate(
            [s.store, s.date, s.product, s.category, s.quantity, s.unit_price, s.revenue], 1
        ):
            cell = ws.cell(row=r, column=col, value=v)
            style_body(cell, num=col >= 5)
            if col >= 6:
                cell.number_format = "¥#,##0"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autofit(ws)


# --------------- CLI ---------------


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", default=str(HERE / "input"))
    parser.add_argument("--output", default=str(HERE / "output" / "monthly_report.xlsx"))
    parser.add_argument("--sheet-name", default="売上明細")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    sales = load_store_excels(input_dir, sheet_name=args.sheet_name)
    print(f"[INFO] {len(sales):,}件の取引を読込")

    agg = aggregate(sales)
    write_report(output_path, sales, agg)

    kpi = agg["kpi"]
    print(f"[OK] レポート生成: {output_path}")
    print(f"     期間: {kpi['first_date']} 〜 {kpi['last_date']}")
    print(f"     店舗数: {kpi['store_count']} / 総売上: ¥{kpi['total_revenue']:,}")
    print(f"     取引件数: {kpi['total_count']:,} / 平均単価: ¥{kpi['avg_unit_price']:,}")


if __name__ == "__main__":
    main()
